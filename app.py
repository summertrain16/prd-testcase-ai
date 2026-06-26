import os
import re
import hashlib
from io import BytesIO
import pandas as pd


import streamlit as st
from dotenv import load_dotenv
from openai import OpenAI
from pypdf import PdfReader
from docx import Document
from openpyxl import load_workbook

STEP_INPUT = "1. 输入材料"
STEP_PENDING = "2. 待确认点收敛"
STEP_FINAL = "3. 最终版需求提炼"
STEP_TEST_CASE = "4. 测试用例与 SQL"
STEP_OPTIONS = [
    STEP_INPUT,
    STEP_PENDING,
    STEP_FINAL,
    STEP_TEST_CASE,
]

# =========================
# 1. 加载环境变量（仅作为默认值，实际从侧边栏 session_state 读取）
# =========================

load_dotenv()

_ENV_API_KEY = os.getenv("OPENAI_API_KEY", "")
_ENV_BASE_URL = os.getenv("OPENAI_BASE_URL", "https://api.openai.com/v1")
_ENV_MODEL = os.getenv("OPENAI_MODEL", "gpt-4o-mini")


# =========================
# 2. Prompt：第一步 PRD 初版解析
# =========================

PRD_DRAFT_ANALYSIS_PROMPT = """
你是一名资深数据测试分析专家，擅长从 PRD、Excel 需求文档、会议纪要、表结构、分区信息和开发代码中提炼数据测试需求。

现在你需要对用户提供的信息进行第一轮分析，输出“初版需求提炼”和“待确认点清单”。

重要原则：
1. 你的目标是辅助生成数据测试用例和 SQL 校验脚本，不是做产品评审、元数据治理评审或开发方案评审。
2. 待确认点必须精简，只保留“会直接影响测试用例设计、SQL 校验逻辑、结果表落数正确性”的必要问题。
3. 不要把所有不清楚的背景信息都列为待确认点。
4. 如果某个问题即使不确认，也不影响测试 SQL、测试断言、字段映射、过滤条件、聚合口径、结果表校验，则不要放入待确认点。
5. 如果用户已经在会议纪要、源表表结构、结果表表结构、分区信息或开发代码中提供了相关信息，则不要重复提问。
6. 不要询问字段中文含义是否需要同步维护到数据字典、元数据、指标平台、数据资产平台等治理类问题。
7. 不要询问字段中文名、业务术语解释、页面展示文案，除非该术语直接决定计算口径、过滤条件或字段取值规则。
8. 不要询问“本次数据加工是全量刷新还是增量更新”，除非 PRD 明确要求测试增量链路，或者该问题会直接影响测试 SQL 的取数范围、分区范围或数据断言。
9. 不要询问“目标表是否存在分区字段及写入分区规则”，如果用户已经上传了结果表表结构并填写了分区信息。
10. 不要询问“源表是否存在分区字段及读取分区规则”，如果用户已经上传了源表表结构并填写了分区信息。
11. 待确认点数量应尽量少。优先控制在 3 到 8 条。如果确实没有阻塞测试设计的问题，可以输出“无”。
12. 不要为了凑数量而提出低价值问题。
13. “待确认点清单”的最后一列必须是“用户补充说明”，该列请留空，不要代替用户填写。

只有以下类型的问题才允许进入待确认点清单：
1. 结果表、结果字段、字段映射关系无法确定。
2. 源表、源字段、关联键无法确定。
3. 核心指标或字段的计算公式、取值规则、枚举映射无法确定。
4. 核心过滤条件无法确定，例如订单状态、是否剔除退款、是否排除测试数据、是否只统计有效数据。
5. 聚合维度、统计粒度、去重规则无法确定。
6. 时间范围、业务日期、分区取数规则未提供，且用户没有通过表结构或分区输入补充。
7. 多表关联方式、主外键、关联条件无法确定。
8. 空值、异常值、重复数据处理规则会影响最终结果，但 PRD 未说明。
9. PRD 与会议纪要、表结构、开发代码之间存在直接矛盾，并且该矛盾会影响测试预期结果。

以下类型的问题禁止进入待确认点清单：
1. 数据字典、元数据、字段中文名、指标平台是否同步维护。
2. 需求背景、业务价值、负责人、上线时间、权限、页面文案等不影响数据校验的问题。
3. 技术实现偏好，例如是否全量刷新、是否增量更新、是否用临时表，除非直接影响测试数据范围或校验断言。
4. 已经由用户上传的源表表结构、结果表表结构、分区信息明确提供的问题。
5. 对字段中文含义的泛泛确认，例如“广告配送调整”“广告返点”的中文含义是否需要维护。
6. 不影响结果表数据正确性的非阻塞建议。

请按以下结构输出：

## 一、初版需求提炼表

| 需求编号 | 需求模块/页面 | 需求描述 | 指标/字段 | 业务口径 | 过滤条件 | 分组维度 | 排序规则 | 来源表/字段 | 结果表/字段 | 验收标准 | 不确定点 |
|---|---|---|---|---|---|---|---|---|---|---|---|

## 二、初版数据加工逻辑梳理

请用条目形式描述可能的数据流、计算逻辑、过滤逻辑、聚合逻辑、结果落表逻辑。

## 三、待确认点清单

如果存在会阻塞测试设计或 SQL 校验的问题，请输出表格：

| 待确认编号 | 待确认问题 | 影响范围 | 建议用户补充内容 | 用户补充说明 |
|---|---|---|---|---|

如果不存在会阻塞测试设计或 SQL 校验的问题，请输出：

无
"""


# =========================
# 3. Prompt：第二步 PRD 最终解析
# =========================

PRD_FINAL_ANALYSIS_PROMPT = """
你是一名资深数据测试分析专家。

现在你需要基于以下信息进行第二轮 PRD 分析，并输出“最终版需求提炼表”：

1. PRD 原文
2. 第一轮 AI 生成的初版需求提炼和待确认点
3. 用户针对待确认点补充的说明
4. 会议纪要
5. 源表表结构及分区信息
6. 结果表表结构及分区信息

要求：
1. 必须优先采纳用户对待确认点的补充说明。
2. 如果用户补充说明和 PRD 原文冲突，需要在最终结果中标记“存在冲突，待确认”。
3. 如果待确认点已经被用户补充清楚，需要在最终版中消除对应不确定性。
4. 如果用户仍未补充清楚，不允许脑补，必须保留为“待确认”。
5. 最终结果要能直接用于后续生成数据测试用例和 SQL 校验脚本。
6. 需要重点输出字段、口径、表、过滤条件、分区、聚合逻辑、验收标准、测试关注点。
7. 如果有源表、结果表、字段映射，请明确列出。
8. 如果没有明确源表或结果表，也要明确说明。
9. 如果源表或结果表结构中提供了分区信息，请在最终结果中体现。
10. 如果表结构和 PRD 或用户补充说明存在冲突，需要明确标记。

请按以下结构输出：

## 一、最终版需求提炼表

| 需求编号 | 需求模块/页面 | 需求描述 | 指标/字段 | 最终业务口径 | 过滤条件 | 分组维度 | 排序规则 | 来源表/字段 | 结果表/字段 | 分区要求 | 验收标准 | 测试关注点 | 状态 |
|---|---|---|---|---|---|---|---|---|---|---|---|---|---|

状态可选：
- 已明确
- 部分待确认
- 待确认
- 存在冲突

## 二、最终数据加工逻辑

请按数据流顺序描述：
1. 数据来源
2. 过滤条件
3. 字段映射
4. 指标计算
5. 聚合逻辑
6. 结果落表
7. 分区逻辑

## 三、字段和口径映射表

| 业务字段/指标 | 来源字段 | 结果字段 | 计算口径 | 空值处理 | 精度要求 | 备注 |
|---|---|---|---|---|---|---|

## 四、最终测试关注点

请列出后续生成数据测试用例时必须覆盖的测试点。

## 五、仍需确认的问题

如果仍然有不清楚的地方，请列出。
如果没有，请写“无”。
"""
PRD_FINAL_ANALYSIS_PROMPT += """
补充约束：
1. 最终版需求提炼只关注数据测试用例和 SQL 校验。
2. 不要保留数据字典、元数据、字段中文名维护、业务术语解释类问题。
3. 不要重复询问用户已经通过表结构和分区信息提供的内容。
4. 不要因为“全量刷新还是增量更新”这类实现方式问题阻塞测试设计，除非它直接决定测试 SQL 的取数范围。
5. 对于不影响测试结果的问题，不要写入最终版风险或待确认项。
"""

# =========================
# 3.1 Prompt：多轮待确认点收敛解析
# =========================
PRD_ITERATIVE_PENDING_ANALYSIS_PROMPT = """
你是一名资深数据测试分析专家。
现在你需要基于以下信息进行新一轮需求解析和待确认点收敛：
1. PRD 原文
2. 上一轮 AI 需求分析结果
3. 上一轮待确认点
4. 用户针对待确认点填写的补充说明
5. 历轮待确认点处理记录
6. 会议纪要
7. 源表表结构及分区信息
8. 结果表表结构及分区信息
9. 开发代码
你的目标：
1. 根据用户补充说明，消除已经明确的待确认点。
2. 如果用户补充说明不足，保留仍会影响测试用例设计或 SQL 校验的待确认点。
3. 如果用户补充说明和 PRD、会议纪要、表结构、开发代码存在冲突，需要继续作为待确认点输出。
4. 可以根据新信息发现新的关键待确认点，但必须严格控制数量。
5. 不要重复输出已经明确的问题。
6. 不要输出不影响测试 SQL、字段映射、过滤条件、聚合口径、结果表校验的问题。
7. 如果没有会阻塞测试设计或 SQL 校验的问题，待确认点清单输出“无”。
只有以下类型的问题才允许进入待确认点清单：
1. 结果表、结果字段、字段映射关系无法确定。
2. 源表、源字段、关联键无法确定。
3. 核心指标或字段的计算公式、取值规则、枚举映射无法确定。
4. 核心过滤条件无法确定，例如订单状态、是否剔除退款、是否排除测试数据、是否只统计有效数据。
5. 聚合维度、统计粒度、去重规则无法确定。
6. 时间范围、业务日期、分区取数规则未提供，且用户没有通过表结构或分区输入补充。
7. 多表关联方式、主外键、关联条件无法确定。
8. 空值、异常值、重复数据处理规则会影响最终结果，但 PRD 未说明。
9. PRD、用户补充说明、会议纪要、表结构、开发代码之间存在直接矛盾，并且该矛盾会影响测试预期结果。
以下类型的问题禁止进入待确认点清单：
1. 数据字典、元数据、字段中文名、指标平台是否同步维护。
2. 需求背景、业务价值、负责人、上线时间、权限、页面文案等不影响数据校验的问题。
3. 技术实现偏好，例如是否全量刷新、是否增量更新、是否用临时表，除非直接影响测试数据范围或校验断言。
4. 已经由用户上传的源表表结构、结果表表结构、分区信息明确提供的问题。
5. 对字段中文含义的泛泛确认。
6. 不影响结果表数据正确性的非阻塞建议。
请按以下结构输出：
## 一、本轮需求提炼结果
| 需求编号 | 需求模块/页面 | 需求描述 | 指标/字段 | 当前业务口径 | 过滤条件 | 分组维度 | 排序规则 | 来源表/字段 | 结果表/字段 | 验收标准 | 状态 | 不确定点 |
|---|---|---|---|---|---|---|---|---|---|---|---|---|
状态可选：
- 已明确
- 部分待确认
- 待确认
- 存在冲突
## 二、本轮数据加工逻辑梳理
请用条目形式描述当前已经明确的数据流、计算逻辑、过滤逻辑、聚合逻辑、结果落表逻辑。
## 三、待确认点清单
如果仍存在会影响测试用例设计或 SQL 校验的问题，请输出表格：
| 待确认编号 | 待确认问题 | 影响范围 | 建议用户补充内容 | 用户补充说明 |
|---|---|---|---|---|
要求：
1. “用户补充说明”列必须留空。
2. 不要重复输出已经被用户补充说明解决的问题。
3. 如果用户补充说明不完整，可以继续保留该问题，但需要说明仍缺少什么。
4. 如果用户补充说明存在冲突，需要明确指出冲突点。
5. 待确认点数量应尽量少，优先控制在 3 到 8 条。
6. 如果没有待确认点，请输出：
无
"""

# =========================
# 4. Prompt：生成测试用例
# =========================

TEST_GEN_PROMPT = """
你是一名资深的数据测试专家，专注于 DataWorks / Hive / ODPS SQL 数据测试。

你需要根据用户提供的以下信息生成数据测试用例和 SQL 校验脚本。你的目标不是穷举所有测试点，而是生成“精简、必要、可定位问题、能发现核心数据问题”的测试设计：

1. PRD 原文
2. 最终版需求提炼表
3. 用户针对待确认点补充的说明
4. 会议纪要
5. 源表表结构 / DDL / 字段说明 / 分区信息
6. 结果表表结构 / DDL / 字段说明 / 分区信息
7. 参考开发代码

必须遵守：
1. 必须以“最终版需求提炼表”为核心依据。
2. 如果最终版需求提炼表中存在“待确认”“部分待确认”“存在冲突”的内容，需要在测试用例和待确认问题中明确标记。
3. 不允许根据初版分析或不明确内容自行脑补。
4. 如果提供了参考开发代码，需要结合代码逻辑生成校验 SQL。
5. 如果提供了源表和结果表结构，需要优先基于真实字段生成 SQL。
6. 如果会议纪要、用户补充说明和 PRD 存在冲突，需要在“待确认问题清单”中指出。
7. 如果上传的表结构文件对应的分区信息为“无分区”，生成 SQL 时不要强行添加该表分区条件，但需要结合最终需求判断是否存在分区待确认问题。
8. 非 ads 层表必须带分区：
   pt='YYYYMMDD'
   如果分区字段未知，用【分区字段】占位。
9. SQL 不要使用 SELECT *。
10. 每条测试用例都要包含：测试目标、前置条件、测试步骤、预期结果、校验 SQL 编号。
11. SQL 脚本中的每段 SQL 需要通过注释标明对应的用例编号，方便测试用例与 SQL 一一对应。
12. 所有数值字段建议使用 NVL(字段, 0) 或 COALESCE(字段, 0) 进行空值处理。
13. 数值差异判断统一使用：
    ABS(NVL(a.字段, 0) - NVL(b.字段, 0)) > 0.1
14. 字符串字段差异判断建议使用：
    NVL(a.字段, '') <> NVL(b.字段, '')
15. 日期、时间、状态、枚举类字段需要按照字段实际类型选择合适的空值兜底方式。
16. 不要按照“字段 × 测试类型”机械展开测试用例。
17. 不要生成大量低价值的类型、长度、格式、元数据、权限、页面展示、性能、监控告警类测试，除非最终版需求明确要求且会直接影响数据正确性。

测试重点优先级如下：
1. P0：结果表主键唯一性校验。
2. P0：结果表数据与源表按加工逻辑计算后的结果一致性比对。
3. P1：结果表关键字段非空校验，仅限主键、分区字段、核心金额/数量/状态字段。
4. P1：结果表行数或分区范围合理性校验，仅在需求明确有必要时生成。

字段一致性比对拆分规则：
1. 不要把所有字段的一致性比对都混在同一个测试用例或同一段 SQL 中。
2. 也不要对所有字段无脑逐字段拆分，避免测试用例过多、重复 SQL 过多。
3. 字段一致性比对需要按照以下维度合理拆分：
   - 来源表；
   - 加工逻辑；
   - 关联条件；
   - 过滤条件；
   - 分区条件；
   - 字段复杂度；
   - 差异定位难度。
4. 如果多个字段同时满足以下条件，可以合并到同一个一致性比对用例和同一段 SQL：
   - 来源表相同；
   - 关联键相同；
   - 过滤条件相同；
   - 分区条件相同；
   - 字段都是直接映射、简单取值或简单重命名；
   - 没有复杂 case when、聚合、去重、优先级、金额计算、枚举转换、空值兜底等逻辑。
5. 简单字段合并时，每条一致性比对用例建议最多合并 3 到 5 个字段。
6. 如果字段存在以下任一情况，应单独生成一个一致性比对用例和一段 SQL：
   - 字段来自不同源表；
   - 字段加工逻辑复杂；
   - 字段涉及金额、比例、返点、调整、冲销、退款、优惠、补贴等计算；
   - 字段涉及 case when、枚举映射、状态转换；
   - 字段涉及聚合、去重、排序取最新、优先级取值；
   - 字段涉及空值兜底、默认值、异常值处理；
   - 字段依赖多个中间字段计算；
   - 字段差异需要单独定位原因。
7. 对复杂字段，应优先保证 SQL 清晰、差异易定位，而不是强行合并。
8. 如果多个复杂字段虽然来自同一源表，但加工逻辑不同，也应该拆成多条一致性比对 SQL。
9. 如果多个字段来源表和加工逻辑基本一致，可以合并比对，但 final select 中仍必须分别展示每个字段的 source 值、result 值和 diff_flag。
10. 测试用例数量要适度精简，但不能为了精简牺牲可读性和问题定位效率。

测试用例数量控制规则：
1. 测试用例应保持精简，但不强制压缩到极少数量。
2. 主键唯一性校验通常保留 1 条。
3. 字段一致性比对应根据字段逻辑合理拆分。
4. 如果字段都是简单直接映射，可以合并生成 1 到 2 条一致性比对用例。
5. 如果字段存在复杂加工逻辑，应按字段或字段组拆分生成一致性比对用例。
6. 一般情况下，测试用例数量建议控制在 4 到 10 条。
7. 如果字段较少但每个字段逻辑都复杂，可以一个字段一条一致性比对用例。
8. 如果字段较多但大部分字段是简单同源同逻辑字段，可以适当合并，避免用例数量膨胀。
9. 不要为了凑数量生成低价值测试点。

请按以下结构输出：

## 一、核心测试关注点

只输出最关键的测试关注点，建议 3 到 6 条。

| 序号 | 测试关注点 | 优先级 | 说明 |
|---|---|---|---|

关注点必须优先围绕：
1. 主键唯一性。
2. 源表加工结果与结果表一致性。
3. 复杂字段加工逻辑正确性。
4. 必要非空。
5. 必要的行数或分区范围校验。
6. 待确认或冲突内容对测试的影响。

## 二、测试用例清单

测试用例需要精简但便于定位问题。不要把所有字段的一致性比对都塞进一个超长用例里。

| 用例编号 | 关联需求编号 | 测试点 | 测试类型 | 前置条件 | 测试步骤 | 预期结果 | 优先级 | 校验SQL编号 |
|---|---|---|---|---|---|---|---|---|

用例生成要求：
1. 必须包含主键唯一性校验用例。
2. 必须包含结果表与源表加工结果一致性比对用例。
3. 一致性比对用例需要根据字段复杂度合理拆分：
   - 简单同源、同关联、同过滤、同分区、直接映射字段可以合并；
   - 复杂字段、金额计算字段、枚举映射字段、case when 字段、聚合字段、去重字段、不同源表字段应拆分。
4. 每条测试用例需要能通过“校验SQL编号”对应到第三部分 SQL 脚本中的具体 SQL。
5. 如果某个需求存在待确认或冲突，需要在测试点或前置条件中明确标记“待确认”。

## 三、SQL 校验脚本

请输出可以直接保存为 .sql 文件并在 VSCode 中打开运行的 SQL 脚本。

SQL 必须放在一个或多个 ```sql 代码块中。

SQL 校验脚本要求：
1. SQL 代码块内部只能包含 SQL 语句和 SQL 注释，不要包含 Markdown 表格、Markdown 标题或普通说明文字。
2. 所有说明必须使用 SQL 注释：
   - 单行注释使用 --
   - 多行注释使用 /* ... */
3. 每段 SQL 必须以分号 ; 结尾。
4. SQL 不要使用 SELECT *。
5. SQL 中每段脚本前必须用注释标明：
   - SQL 编号；
   - 对应用例编号；
   - 测试目标；
   - 适用字段；
   - 是否存在待确认事项。
6. 主键唯一性 SQL 应输出重复主键明细，不要只输出重复数量。
7. 每个一致性比对 SQL 建议采用 expected / actual CTE 结构：
   - expected：源表按需求逻辑加工后的期望结果；
   - actual：结果表实际落表结果；
   - final select：对 expected 和 actual 进行 full outer join、full join 或 left join 比对。
8. final select 必须同时输出：
   - 主键字段；
   - 必要的分区字段；
   - 源表加工后的期望字段值，命名为 source_xxx 或 expected_xxx；
   - 结果表实际字段值，命名为 result_xxx 或 actual_xxx；
   - 字段差异标识，命名为 xxx_diff_flag；
   - 差异类型，命名为 diff_type；
   - 如果字段逻辑复杂，需要输出中间计算字段，方便定位差异原因。
9. final select 不允许只输出 count 数量。必须输出差异明细，方便定位具体哪条数据、哪个字段不一致。
10. 如果多个简单字段合并比对，final select 中也必须分别展示每个字段的：
    - source_xxx 或 expected_xxx；
    - result_xxx 或 actual_xxx；
    - xxx_diff_flag。
11. 如果是复杂字段单独比对，expected CTE 中必须尽量保留计算该字段所需的中间字段，例如原始金额、调整金额、退款金额、状态字段、枚举字段等。
12. where 条件中只筛选存在差异的数据，例如：
    - 源表有但结果表无；
    - 结果表有但源表无；
    - 源表加工值与结果表值不一致。
13. diff_type 建议按照以下规则输出：
    - SOURCE_ONLY：源表加工后存在，但结果表不存在；
    - RESULT_ONLY：结果表存在，但源表加工后不存在；
    - FIELD_VALUE_DIFF：主键存在但字段值不一致；
    - MULTI_FIELD_DIFF：多个字段同时不一致。
14. 对金额、数量、比例等数值类字段，差异判断统一使用：
    ABS(NVL(expected_字段, 0) - NVL(result_字段, 0)) > 0.1
15. 对字符串类字段，差异判断使用：
    NVL(expected_字段, '') <> NVL(result_字段, '')
16. 对日期、时间字段，需根据实际字段类型选择合适的比较方式；如果无法确认类型，需要用 SQL 注释标记 TODO。
17. 如果存在小数精度问题，可以使用 ROUND 或 ABS 差值阈值比较，优先使用 ABS 差值阈值。
18. 如果表名、字段名、分区值已提供，请直接使用真实名称。
19. 如果缺少必要信息，可以用 SQL 注释标注 TODO，但不要让 TODO 文字破坏 SQL 文件可执行性。
20. 对于非 ads 层表，如果有分区信息，必须添加分区条件；如果分区字段未知，用【分区字段】占位。
21. 对于明确“无分区”的表，不要强行添加分区条件。
22. 如果 SQL 方言存在差异，优先使用 DataWorks / Hive / ODPS 兼容写法。
23.若无明确要求，则都是结果表左关联源表

SQL 示例格式如下，生成时请参考该结构，但必须使用用户提供的真实表名和字段名：

```sql
/* =====================================================
 SQL-001
 对应用例编号：TC-001
 测试目标：结果表主键唯一性校验
 适用字段：主键字段
 待确认事项：无
 ===================================================== */

SELECT
    id,
    pt,
    COUNT(1) AS duplicate_cnt
FROM ads_result_table
WHERE pt = 'YYYYMMDD'
GROUP BY
    id,
    pt
HAVING COUNT(1) > 1;


/* =====================================================
 SQL-002
 对应用例编号：TC-002
 测试目标：简单同源字段一致性比对
 适用字段：field_a, field_b
 待确认事项：无
 ===================================================== */

WITH expected AS (
    SELECT
        id,
        pt,
        field_a AS expected_field_a,
        field_b AS expected_field_b
    FROM dwd_source_table
    WHERE pt = 'YYYYMMDD'
),
actual AS (
    SELECT
        id,
        pt,
        field_a AS result_field_a,
        field_b AS result_field_b
    FROM ads_result_table
    WHERE pt = 'YYYYMMDD'
)
SELECT
    NVL(e.id, a.id) AS id,
    NVL(e.pt, a.pt) AS pt,

    e.expected_field_a,
    a.result_field_a,
    e.expected_field_b,
    a.result_field_b
FROM actual a
left join expected e
    ON e.id = a.id
   AND e.pt = a.pt
WHERE
     NVL(e.expected_field_a, '') <> NVL(a.result_field_a, '')
    OR NVL(e.expected_field_b, '') <> NVL(a.result_field_b, '');


/* =====================================================
 SQL-003
 对应用例编号：TC-003
 测试目标：复杂金额字段一致性比对
 适用字段：amount_field
 待确认事项：如金额计算口径未明确，请标记 TODO
 ===================================================== */

WITH source_base AS (
    SELECT
        id,
        pt,
        original_amt,
        adjust_amt,
        refund_amt,
        status
    FROM dwd_source_table
    WHERE pt = 'YYYYMMDD'
),
expected AS (
    SELECT
        id,
        pt,

        -- 保留中间字段，方便定位复杂字段差异原因
        original_amt AS source_original_amt,
        adjust_amt AS source_adjust_amt,
        refund_amt AS source_refund_amt,
        status AS source_status,

        NVL(original_amt, 0)
        + NVL(adjust_amt, 0)
        - NVL(refund_amt, 0) AS expected_amount_field
    FROM source_base
),
actual AS (
    SELECT
        id,
        pt,
        amount_field AS result_amount_field
    FROM ads_result_table
    WHERE pt = 'YYYYMMDD'
)
SELECT
    NVL(e.id, a.id) AS id,
    NVL(e.pt, a.pt) AS pt,

    e.source_original_amt,
    e.source_adjust_amt,
    e.source_refund_amt,
    e.source_status,

    e.expected_amount_field,
    a.result_amount_field
FROM actual a
left join expected e
    ON e.id = a.id
   AND e.pt = a.pt
WHERE
    ABS(NVL(e.expected_amount_field, 0) - NVL(a.result_amount_field, 0)) > 0.1;

```
## 四、待确认问题清单

列出仍然无法生成准确 SQL 或准确测试预期的问题。
"""


# =========================
# 5. LLM 调用方法
# =========================

def call_llm(system_prompt: str, user_content: str) -> str:
    """
    调用大模型，返回文本结果。
    兼容 OpenAI SDK 标准对象、dict、字符串返回。
    从 session_state 读取 API Key / Base URL / Model，支持每个用户自己配置。
    """
    api_key = st.session_state.get("user_openai_api_key", "").strip()
    base_url = st.session_state.get("user_openai_base_url", "").strip()
    model = st.session_state.get("user_openai_model", "").strip()

    if not api_key:
        return "错误：未配置 API Key，请在左侧侧边栏填写你的 API Key。"

    if not base_url:
        base_url = "https://api.openai.com/v1"
    if not model:
        model = "gpt-4o-mini"

    try:
        llm_client = OpenAI(
            api_key=api_key,
            base_url=base_url
        )

        response = llm_client.chat.completions.create(
            model=model,
            messages=[
                {
                    "role": "system",
                    "content": system_prompt
                },
                {
                    "role": "user",
                    "content": user_content
                }
            ],
            temperature=0.2
        )

        if isinstance(response, str):
            if response.strip().lower().startswith("<!doctype html") or "<html" in response.lower():
                return (
                    "调用大模型失败：接口返回了 HTML 页面，而不是模型结果。\n\n"
                    "这通常说明 Base URL 配置错误。\n"
                    "如果你使用 New API，请把 Base URL 改成：\n\n"
                    "http://你的NewAPI地址/v1\n\n"
                    "注意不要写成后台首页地址，也不要写成 /v1/chat/completions。"
                )
            return response

        if isinstance(response, dict):
            try:
                return response["choices"][0]["message"]["content"]
            except Exception:
                return str(response)

        if hasattr(response, "choices"):
            return response.choices[0].message.content

        return str(response)

    except Exception as e:
        return f"调用大模型失败：{str(e)}"


def is_llm_error(result: str) -> bool:
    """
    判断大模型返回是否为错误信息。
    """
    if not result:
        return True

    text = str(result).strip()

    return (
        text.startswith("错误：")
        or text.startswith("调用大模型失败：")
    )

# =========================
# 6. 文件读取方法
# =========================

def read_xlsx_file(uploaded_file) -> str:
    """
    读取普通 xlsx 文件内容，并转换成文本，方便大模型分析。
    用于 PRD Excel 文件读取。
    """
    try:
        uploaded_file.seek(0)

        workbook = load_workbook(
            uploaded_file,
            data_only=True,
            read_only=True
        )

        result = []

        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            result.append(f"\n\n## Excel Sheet：{sheet_name}\n")

            rows = list(sheet.iter_rows(values_only=True))

            if not rows:
                result.append("该 Sheet 为空。\n")
                continue

            max_rows = 300
            rows = rows[:max_rows]

            for row_index, row in enumerate(rows, start=1):
                values = []
                for cell in row:
                    if cell is None:
                        values.append("")
                    else:
                        values.append(str(cell).strip())
                result.append(f"第 {row_index} 行：" + " | ".join(values))

            if sheet.max_row > max_rows:
                result.append(
                    f"\n注意：该 Sheet 共 {sheet.max_row} 行，当前只读取前 {max_rows} 行。"
                )

        return "\n".join(result)

    except Exception as e:
        return f"xlsx 文件读取失败：{str(e)}"


def read_uploaded_file(uploaded_file) -> str:
    """
    支持读取 txt、md、sql、csv、json、py、pdf、docx、xlsx。
    用于 PRD 文件读取。
    """
    if uploaded_file is None:
        return ""

    file_name = uploaded_file.name.lower()

    try:
        if file_name.endswith((".txt", ".md", ".sql", ".csv", ".json", ".py")):
            uploaded_file.seek(0)
            return uploaded_file.read().decode("utf-8", errors="ignore")

        elif file_name.endswith(".pdf"):
            uploaded_file.seek(0)
            reader = PdfReader(uploaded_file)
            text = ""
            for page in reader.pages:
                page_text = page.extract_text()
                if page_text:
                    text += page_text + "\n"
            return text

        elif file_name.endswith(".docx"):
            uploaded_file.seek(0)
            doc = Document(uploaded_file)
            text = ""
            for para in doc.paragraphs:
                text += para.text + "\n"
            return text

        elif file_name.endswith(".xlsx"):
            return read_xlsx_file(uploaded_file)

        else:
            return "暂不支持该文件格式，请上传 txt、md、sql、csv、json、py、pdf、docx 或 xlsx 文件。"

    except Exception as e:
        return f"文件读取失败：{str(e)}"


def read_table_schema_xlsx(file_bytes: bytes, file_name: str) -> str:
    """
    读取表结构 xlsx 文件内容，转换成文本。
    一个 xlsx 可以包含多个 sheet。
    用于源表/结果表结构上传。
    """
    try:
        workbook = load_workbook(
            BytesIO(file_bytes),
            data_only=True,
            read_only=True
        )

        result = []

        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]

            result.append(f"\n## 文件：{file_name}")
            result.append(f"## Sheet：{sheet_name}")

            rows = []

            for row in sheet.iter_rows(values_only=True):
                values = []
                has_value = False

                for cell in row:
                    if cell is None:
                        values.append("")
                    else:
                        cell_value = str(cell).strip()
                        values.append(cell_value)
                        if cell_value:
                            has_value = True

                if has_value:
                    rows.append(values)

            if not rows:
                result.append("该 Sheet 为空。")
                continue

            max_rows = 1000
            original_row_count = len(rows)
            rows = rows[:max_rows]

            for index, row in enumerate(rows, start=1):
                result.append(f"第 {index} 行：" + " | ".join(row))

            if original_row_count > max_rows:
                result.append(
                    f"注意：该 Sheet 共 {original_row_count} 行，当前只读取前 {max_rows} 行。"
                )

        return "\n".join(result)

    except Exception as e:
        return f"xlsx 表结构文件读取失败：{str(e)}"


def get_uploaded_file_id(file_name: str, file_bytes: bytes) -> str:
    """
    根据文件名和文件内容生成唯一 ID，用于去重和删除。
    """
    md5 = hashlib.md5()
    md5.update(file_name.encode("utf-8"))
    md5.update(file_bytes)
    return md5.hexdigest()

PENDING_POINT_COLUMNS = [
    "待确认编号",
    "待确认问题",
    "影响范围",
    "建议用户补充内容",
    "用户补充说明"
]


def is_markdown_separator_row(cells: list[str]) -> bool:
    """
    判断 Markdown 表格中的分隔行，例如：
    |---|---|---|
    """
    if not cells:
        return False

    for cell in cells:
        cell = cell.strip()
        if not re.fullmatch(r":?-{3,}:?", cell):
            return False

    return True

def go_to_step(step_name: str) -> None:
    st.session_state["current_step"] = step_name
    st.rerun()

def on_step_radio_change() -> None:
    st.session_state["current_step"] = st.session_state["current_step_radio"]

def get_materials_from_state() -> dict:
    return {
        "prd_text": st.session_state.get("prd_text", ""),
        "meeting_notes": st.session_state.get("meeting_notes", ""),
        "result_table_schema": st.session_state.get("result_table_schema", ""),
        "source_table_schema": st.session_state.get("source_table_schema", ""),
        "dev_code": st.session_state.get("dev_code", ""),
    }

def parse_pending_points_from_markdown(markdown_text: str) -> list[dict]:
    """
    从第一步 AI 输出的 Markdown 中解析“待确认点清单”表格。
    兼容：
    - ## 三、待确认点清单
    - ### 三、待确认点清单
    - 三、待确认点清单
    - ## 三、待确认问题清单
    - ## 三、待确认事项清单
    """
    if not markdown_text:
        return []

    lines = markdown_text.splitlines()

    # 1. 找到待确认点章节起始行，不强依赖“## 三、待确认点清单”
    start_idx = None

    for i, line in enumerate(lines):
        clean_line = line.strip()

        if "待确认" in clean_line and (
            "清单" in clean_line or "问题" in clean_line or "事项" in clean_line
        ):
            start_idx = i
            break

    if start_idx is None:
        return []

    # 2. 找到下一个章节标题作为结束位置
    end_idx = len(lines)

    for j in range(start_idx + 1, len(lines)):
        clean_line = lines[j].strip()

        # 匹配类似：## 四、xxx / ### 四、xxx / 四、xxx
        if re.match(r"^\s*#{0,6}\s*[一二三四五六七八九十]+[、.．]\s*", clean_line):
            if "待确认" not in clean_line:
                end_idx = j
                break

    section_text = "\n".join(lines[start_idx:end_idx])

    # 3. 如果明确是无，则返回空
    if re.search(r"(?m)^\s*无\s*$", section_text.strip()):
        return []

    # 4. 提取 Markdown 表格行
    table_lines = []

    for line in section_text.splitlines():
        line = line.strip()
        if line.startswith("|") and "|" in line:
            table_lines.append(line)

    if not table_lines:
        return []

    header = None
    rows = []

    for line in table_lines:
        cells = [cell.strip() for cell in line.strip().strip("|").split("|")]

        if is_markdown_separator_row(cells):
            continue

        if header is None:
            header = cells
            continue

        if not cells:
            continue

        if len(cells) < len(header):
            cells = cells + [""] * (len(header) - len(cells))

        row_dict = {}

        for col in PENDING_POINT_COLUMNS:
            if col in header:
                idx = header.index(col)
                row_dict[col] = cells[idx] if idx < len(cells) else ""
            else:
                row_dict[col] = ""

        # 兼容列名变化
        if not row_dict["待确认编号"]:
            for alt_col in ["编号", "问题编号", "待确认点编号"]:
                if alt_col in header:
                    idx = header.index(alt_col)
                    row_dict["待确认编号"] = cells[idx] if idx < len(cells) else ""
                    break

        if not row_dict["待确认问题"]:
            for alt_col in ["待确认点", "问题", "确认问题", "待确认事项"]:
                if alt_col in header:
                    idx = header.index(alt_col)
                    row_dict["待确认问题"] = cells[idx] if idx < len(cells) else ""
                    break

        if not row_dict["影响范围"]:
            for alt_col in ["影响", "影响说明"]:
                if alt_col in header:
                    idx = header.index(alt_col)
                    row_dict["影响范围"] = cells[idx] if idx < len(cells) else ""
                    break

        if not row_dict["建议用户补充内容"]:
            for alt_col in ["建议补充内容", "需补充内容", "用户需补充内容"]:
                if alt_col in header:
                    idx = header.index(alt_col)
                    row_dict["建议用户补充内容"] = cells[idx] if idx < len(cells) else ""
                    break

        # 用户补充说明保持空，等待用户填写
        row_dict["用户补充说明"] = ""

        # 过滤空行
        if any(str(row_dict[col]).strip() for col in PENDING_POINT_COLUMNS[:-1]):
            rows.append(row_dict)

    return rows


def remove_pending_points_section(markdown_text: str) -> str:
    """
    从 AI 输出中移除“待确认点清单”部分。
    页面上改用可编辑表格展示待确认点，避免重复显示不可编辑 Markdown 表格。
    兼容：
    - ## 三、待确认点清单
    - ### 三、待确认问题清单
    - 三、待确认事项
    """
    if not markdown_text:
        return ""

    lines = markdown_text.splitlines()

    start_idx = None

    for i, line in enumerate(lines):
        clean_line = line.strip()
        if "待确认" in clean_line and (
            "清单" in clean_line or "问题" in clean_line or "事项" in clean_line
        ):
            start_idx = i
            break

    if start_idx is None:
        return markdown_text.strip()

    end_idx = len(lines)

    for j in range(start_idx + 1, len(lines)):
        clean_line = lines[j].strip()

        if re.match(r"^\s*#{0,6}\s*[一二三四五六七八九十]+[、.．]\s*", clean_line):
            if "待确认" not in clean_line:
                end_idx = j
                break

    new_lines = lines[:start_idx] + lines[end_idx:]

    return "\n".join(new_lines).strip()


def pending_points_to_llm_text(rows: list[dict]) -> str:
    """
    将可编辑待确认点表格转换成传给第二步大模型的文本。
    """
    if not rows:
        return "未解析到待确认点清单，或第一步未产生待确认点。"

    lines = []

    for row in rows:
        pending_id = str(row.get("待确认编号", "")).strip()
        question = str(row.get("待确认问题", "")).strip()
        impact = str(row.get("影响范围", "")).strip()
        suggestion = str(row.get("建议用户补充内容", "")).strip()
        answer = str(row.get("用户补充说明", "")).strip()

        if not pending_id and not question:
            continue

        lines.append(
            f"""
待确认编号：{pending_id if pending_id else "未提供"}
待确认问题：{question if question else "未提供"}
影响范围：{impact if impact else "未提供"}
建议用户补充内容：{suggestion if suggestion else "未提供"}
用户补充说明：{answer if answer else "未补充"}
"""
        )

    return "\n".join(lines).strip() if lines else "无待确认点。"


def pending_points_to_markdown(rows: list[dict]) -> str:
    """
    将可编辑后的待确认点清单转换成 Markdown，主要用于下载文件。
    """
    if not rows:
        return "## 三、待确认点清单\n\n无"

    lines = []
    lines.append("## 三、待确认点清单")
    lines.append("")
    lines.append("| 待确认编号 | 待确认问题 | 影响范围 | 建议用户补充内容 | 用户补充说明 |")
    lines.append("|---|---|---|---|---|")

    for row in rows:
        line = "| {id} | {question} | {impact} | {suggestion} | {answer} |".format(
            id=str(row.get("待确认编号", "")).replace("|", "\\|"),
            question=str(row.get("待确认问题", "")).replace("|", "\\|"),
            impact=str(row.get("影响范围", "")).replace("|", "\\|"),
            suggestion=str(row.get("建议用户补充内容", "")).replace("|", "\\|"),
            answer=str(row.get("用户补充说明", "")).replace("|", "\\|"),
        )
        lines.append(line)

    return "\n".join(lines)

def get_pending_answer_key(row: dict, index: int) -> str:
    """
    为每个待确认点的“用户补充说明”生成稳定的 widget key。
    key 中包含 pending_points_editor_version，重新生成第一步后会自动换一批 key，避免旧输入串到新结果。
    """
    version = st.session_state.get("pending_points_editor_version", 0)

    raw = (
        f"{version}_"
        f"{index}_"
        f"{row.get('待确认编号', '')}_"
        f"{row.get('待确认问题', '')}"
    )

    md5 = hashlib.md5()
    md5.update(raw.encode("utf-8"))

    return f"pending_answer_{md5.hexdigest()}"

def extract_sql_section_from_test_result(result_text: str) -> str:
    """
    从测试用例结果中提取“三、SQL 校验脚本”章节。
    """
    if not result_text:
        return ""

    pattern = r"(?m)^#{1,6}\s*三[、.．]\s*SQL\s*校验脚本\s*$"
    match = re.search(pattern, result_text)

    if not match:
        return result_text.strip()

    start = match.end()

    # 找下一个同级或任意后续章节标题，例如 ## 四、说明
    next_match = re.search(
        r"(?m)^#{1,6}\s*[四五六七八九十][、.．]\s*",
        result_text[start:]
    )

    if next_match:
        end = start + next_match.start()
        return result_text[start:end].strip()

    return result_text[start:].strip()


def extract_sql_code_blocks(markdown_text: str) -> list[str]:
    """
    提取 Markdown 中的 sql 代码块。
    """
    if not markdown_text:
        return []

    blocks = re.findall(
        r"```(?:sql|SQL)?\s*\n(.*?)```",
        markdown_text,
        flags=re.DOTALL
    )

    return [block.strip() for block in blocks if block.strip()]


def strip_markdown_fence(text: str) -> str:
    """
    去除 Markdown 代码块标记。
    """
    if not text:
        return ""

    text = re.sub(r"```(?:sql|SQL)?", "", text)
    text = text.replace("```", "")

    return text.strip()


def escape_sql_block_comment(text: str) -> str:
    """
    避免测试用例说明中的 */ 破坏 SQL 多行注释。
    """
    if not text:
        return ""

    return text.replace("*/", "* /")

def inject_custom_css() -> None:
    """
    注入页面美化 CSS — 现代极简风格。
    """
    st.markdown(
        """
<style>
/* ===== 全局 ===== */
.block-container {
    padding-top: 2rem;
    padding-bottom: 4rem;
    max-width: 1280px;
}

/* 字体层级收紧 */
h1, h2, h3 {
    font-weight: 600 !important;
    letter-spacing: -0.02em;
}

/* ===== 顶部标题区 ===== */
.app-hero {
    padding: 28px 32px;
    border-radius: 10px;
    background: #2563EB;
    margin-bottom: 24px;
}

.app-hero-title {
    font-size: 22px;
    font-weight: 600;
    color: #F8FAFC;
    margin-bottom: 6px;
    letter-spacing: -0.02em;
}

.app-hero-desc {
    font-size: 13px;
    color: #94A3B8;
    line-height: 1.6;
}

/* ===== 步骤进度卡片 ===== */
.step-card {
    padding: 16px 14px;
    border-radius: 8px;
    border: 1px solid #E2E8F0;
    background: #FFFFFF;
    text-align: center;
    min-height: 72px;
    transition: all 0.15s ease;
}

.step-card-active {
    border: 1px solid #2563EB;
    background: #2563EB;
}

.step-card-done {
    border: 1px solid #E2E8F0;
    background: #F8FAFC;
}

.step-card-title {
    font-size: 13px;
    font-weight: 500;
    color: #334155;
    line-height: 1.4;
}

.step-card-active .step-card-title {
    color: #F8FAFC;
}

.step-card-status {
    margin-top: 6px;
    font-size: 11px;
    color: #94A3B8;
    font-weight: 400;
}

.step-card-active .step-card-status {
    color: #94A3B8;
}

.step-card-done .step-card-status {
    color: #64748B;
}

/* ===== 步骤说明卡片 ===== */
.page-section-card {
    padding: 20px 24px;
    border-radius: 8px;
    background: #FFFFFF;
    border: 1px solid #E2E8F0;
    margin-bottom: 20px;
}

.page-section-title {
    font-size: 18px;
    font-weight: 600;
    color: #0F172A;
    margin-bottom: 4px;
    letter-spacing: -0.01em;
}

.page-section-desc {
    font-size: 13px;
    color: #64748B;
    line-height: 1.6;
}

/* ===== Streamlit 按钮 ===== */
.stButton > button,
.stDownloadButton > button {
    border-radius: 6px !important;
    min-height: 38px;
    font-weight: 500 !important;
    font-size: 13px !important;
    transition: all 0.12s ease;
    border: 1px solid #DBEAFE !important;
    background: #FFFFFF !important;
    color: #2563EB !important;
}

.stButton > button:hover,
.stDownloadButton > button:hover {
    border-color: #2563EB !important;
    background: #EFF6FF !important;
}

/* primary 按钮 */
.stButton > button[kind="primary"],
div[data-testid="stButton"] > button[kind="primary"] {
    background: #2563EB !important;
    color: #FFFFFF !important;
    border: 1px solid #2563EB !important;
}

.stButton > button[kind="primary"]:hover {
    background: #1D4ED8 !important;
    border-color: #1D4ED8 !important;
}

/* ===== 输入框 ===== */
textarea, input {
    border-radius: 6px !important;
    font-size: 13px !important;
}

/* ===== expander ===== */
div[data-testid="stExpander"] {
    border-radius: 8px !important;
    border: 1px solid #E2E8F0 !important;
}

/* ===== sidebar ===== */
section[data-testid="stSidebar"] {
    background: #FAFAFA;
    border-right: 1px solid #E2E8F0;
}

section[data-testid="stSidebar"] h1,
section[data-testid="stSidebar"] h2,
section[data-testid="stSidebar"] h3 {
    color: #0F172A;
    font-size: 13px !important;
    font-weight: 600 !important;
    text-transform: uppercase;
    letter-spacing: 0.05em;
}

/* ===== data editor / dataframe ===== */
div[data-testid="stDataFrame"] {
    border-radius: 8px;
}

/* ===== 分割线 ===== */
hr {
    border: none;
    border-top: 1px solid #F1F5F9;
    margin-top: 1.5rem;
    margin-bottom: 1.5rem;
}

/* ===== alert 框 ===== */
div[data-testid="stAlert"] {
    border-radius: 6px;
    border: 1px solid #E2E8F0;
}

/* ===== 代码块 ===== */
pre {
    border-radius: 6px !important;
}

/* ===== markdown 表格 ===== */
table {
    border-radius: 6px;
    overflow: hidden;
}

th {
    font-weight: 600 !important;
    font-size: 12px !important;
}

td {
    font-size: 13px !important;
}

/* ===== tab ===== */
div[data-testid="stTabs"] {
    border-radius: 8px;
}

/* 隐藏 Streamlit 默认的 padding 减少空隙感 */
.element-container {
    margin-bottom: 12px;
}
</style>
        """,
        unsafe_allow_html=True
    )

def render_markdown_in_scroll_box(
    title: str,
    markdown_text: str,
    height: int = 520,
    expanded: bool = True
) -> None:
    """
    用 expander + 固定高度容器展示长 Markdown，避免页面过长。
    """
    if not markdown_text:
        return

    with st.expander(title, expanded=expanded):
        with st.container(height=height, border=True):
            st.markdown(markdown_text)

def render_step_progress() -> None:
    """
    顶部流程进度展示。
    """
    current_step = st.session_state.get("current_step", STEP_INPUT)

    if current_step in STEP_OPTIONS:
        current_index = STEP_OPTIONS.index(current_step)
    else:
        current_index = 0

    progress_value = (current_index + 1) / len(STEP_OPTIONS)

    st.progress(progress_value)

    cols = st.columns(len(STEP_OPTIONS))

    icons = ["", "", "", ""]

    for index, step in enumerate(STEP_OPTIONS):
        if index < current_index:
            status = "已完成"
            card_class = "step-card step-card-done"
            status_icon = "—"
        elif index == current_index:
            status = "当前步骤"
            card_class = "step-card step-card-active"
            status_icon = "●"
        else:
            status = "待处理"
            card_class = "step-card"
            status_icon = "○"

        with cols[index]:
            st.markdown(
                f"""
<div class="{card_class}">
    <div class="step-card-title">{step}</div>
    <div class="step-card-status">{status_icon} {status}</div>
</div>
                """,
                unsafe_allow_html=True
            )

    st.write("")

def render_page_header(title: str, desc: str, icon: str = "") -> None:
    """
    每个步骤顶部的说明卡片。
    """
    title_html = f"{icon} {title}" if icon else title
    st.markdown(
        f"""
<div class="page-section-card">
    <div class="page-section-title">{title_html}</div>
    <div class="page-section-desc">{desc}</div>
</div>
        """,
        unsafe_allow_html=True
    )

def build_vscode_runnable_sql_download_content(result_text: str) -> str:
    """
    构造可在 VSCode 中打开并执行的 .sql 文件内容。

    文件结构：
    1. 前半部分：测试用例和关注点，以 SQL 多行注释形式保留。
    2. 后半部分：真正可执行 SQL。
    """
    if not result_text:
        return ""

    # 找到 SQL 章节位置
    sql_heading_pattern = r"(?m)^#{1,6}\s*三[、.．]\s*SQL\s*校验脚本\s*$"
    match = re.search(sql_heading_pattern, result_text)

    if match:
        before_sql_section = result_text[:match.start()].strip()
        sql_section = extract_sql_section_from_test_result(result_text)
    else:
        before_sql_section = ""
        sql_section = result_text.strip()

    # 优先提取 ```sql ... ``` 代码块
    sql_blocks = extract_sql_code_blocks(sql_section)

    if sql_blocks:
        sql_body = "\n\n\n".join(sql_blocks).strip()
    else:
        # 兜底：如果模型没有输出代码块，则去掉 Markdown fence 后直接作为 SQL 内容
        sql_body = strip_markdown_fence(sql_section)

    before_sql_section = escape_sql_block_comment(before_sql_section)

    sql_file_content = f"""/*
=====================================================
数据测试用例说明
=====================================================

以下内容来自 AI 生成的测试关注点和测试用例。
为了保证 .sql 文件可以在 VSCode / SQL 客户端中直接打开运行，
测试用例说明已放入 SQL 注释中。

{before_sql_section if before_sql_section else "无额外测试用例说明。"}

=====================================================
*/

-- =====================================================
-- 数据测试 SQL 校验脚本
-- 使用说明：
-- 1. 请确认 VSCode 已安装对应数据库插件，例如 SQLTools、Hive、Spark SQL、Presto、Trino、MySQL 等。
-- 2. 请确认当前连接的数据源与 SQL 方言一致。
-- 3. 如果 SQL 中存在 TODO 注释，请先替换为真实表名、字段名或分区值。
-- 4. 每段 SQL 可单独选中执行，也可以整体执行。
-- =====================================================

{sql_body.rstrip()}

"""
    return sql_file_content

def _split_sections_by_heading(result_text: str, headings: list) -> dict:
    """
    按 heading 关键词将 result_text 拆分为多段。
    headings 示例: ["一", "二", "三"]
    返回: {"一": "...", "二": "...", "三": "..."}
    每个 value 包含从该标题到下一个标题前的完整内容（含标题行）。
    """
    # 构建正则：匹配 ## 一、 或 ## 一. 或 ## 一． 等
    patterns = {}
    for h in headings:
        patterns[h] = re.compile(
            rf"(?m)^(#{{1,6}})\s*{h}[、.．]"
        )

    # 找到每个标题的位置
    positions = []
    for h, p in patterns.items():
        m = p.search(result_text)
        if m:
            positions.append((m.start(), h, m))
    positions.sort(key=lambda x: x[0])

    sections = {}
    for i, (start, h, m) in enumerate(positions):
        end = positions[i + 1][0] if i + 1 < len(positions) else len(result_text)
        sections[h] = result_text[start:end].strip()

    return sections


def render_test_case_result_with_download(result_text: str) -> None:
    """
    渲染测试用例结果：
    - SQL 下载按钮放在最顶部
    - 一、二、三三个章节各自包在可滚动容器里（st.container border + height）
    - 章节之间用 st.divider() 分隔
    """
    if not result_text:
        return

    sql_download_content = build_vscode_runnable_sql_download_content(result_text)

    sections = _split_sections_by_heading(result_text, ["一", "二", "三"])

    # 下载按钮始终放最顶部
    st.download_button(
        label="下载 SQL 脚本",
        data=sql_download_content,
        file_name="data_test_validation.sql",
        mime="text/plain",
        use_container_width=True
    )

    # 如果没匹配到任何章节标题，退回整段渲染（放在可滚动容器里）
    if not sections:
        with st.container(height=400, border=True):
            st.markdown(result_text)
        return

    # 一、测试关注点 — 可滚动容器
    if "一" in sections:
        with st.container(height=350, border=True):
            st.markdown(sections["一"])

    st.divider()

    # 二、测试用例清单 — 可滚动容器
    if "二" in sections:
        with st.container(height=400, border=True):
            st.markdown(sections["二"])

    st.divider()

    # 三、SQL 校验脚本 — 可滚动容器
    if "三" in sections:
        with st.container(height=450, border=True):
            st.markdown(sections["三"])

def render_pending_points_data_editor() -> None:
    """
    使用 st.data_editor 紧凑展示待确认点，只允许编辑“用户补充说明”列。
    """
    rows = st.session_state.get("pending_points_rows", [])

    if not rows:
        st.success("当前没有阻塞测试设计或 SQL 校验的待确认点，可以继续生成最终版需求提炼表。")
        st.session_state["prd_pending_answers"] = "无待确认点。"
        return

    df = pd.DataFrame(rows)

    for col in PENDING_POINT_COLUMNS:
        if col not in df.columns:
            df[col] = ""

    df = df[PENDING_POINT_COLUMNS].fillna("")

    editor_key = f"pending_points_data_editor_{st.session_state.get('pending_points_editor_version', 0)}"

    edited_df = st.data_editor(
        df,
        key=editor_key,
        hide_index=True,
        use_container_width=True,
        height=min(460, 110 + len(df) * 58),
        disabled=[
            "待确认编号",
            "待确认问题",
            "影响范围",
            "建议用户补充内容"
        ],
        column_config={
            "待确认编号": st.column_config.TextColumn(
                "待确认编号",
                width="small"
            ),
            "待确认问题": st.column_config.TextColumn(
                "待确认问题",
                width="large"
            ),
            "影响范围": st.column_config.TextColumn(
                "影响范围",
                width="medium"
            ),
            "建议用户补充内容": st.column_config.TextColumn(
                "建议用户补充内容",
                width="large"
            ),
            "用户补充说明": st.column_config.TextColumn(
                "用户补充说明，可编辑",
                width="large",
                help="请在这里填写确认结果；如果认为可忽略，也可以不填并点击忽略。"
            ),
        },
        num_rows="fixed"
    )

    st.session_state["pending_points_rows"] = edited_df.fillna("").to_dict("records")

    st.session_state["prd_pending_answers"] = pending_points_to_llm_text(
        st.session_state["pending_points_rows"]
    )

def sync_pending_points_from_widgets() -> None:
    """
    从每个待确认点对应的 text_area widget 中同步用户补充说明，
    并更新 prd_pending_answers，供第二步和测试用例生成使用。
    """
    rows = st.session_state.get("pending_points_rows", [])

    if not rows:
        st.session_state["prd_pending_answers"] = "无待确认点。"
        return

    for index, row in enumerate(rows):
        key = get_pending_answer_key(row, index)

        if key in st.session_state:
            row["用户补充说明"] = str(st.session_state[key]).strip()
        else:
            row["用户补充说明"] = str(row.get("用户补充说明", "")).strip()

    st.session_state["pending_points_rows"] = rows
    st.session_state["prd_pending_answers"] = pending_points_to_llm_text(rows)

def render_table_schema_uploader(title: str, state_prefix: str) -> str:
    """
    渲染表结构 xlsx 上传组件。

    参数：
    title: 页面显示名称，例如：结果表表结构、源表表结构
    state_prefix: session_state 前缀，例如：result_schema、source_schema

    返回：
    拼接后的表结构文本，用于传给大模型。
    """
    items_key = f"{state_prefix}_items"
    uploader_version_key = f"{state_prefix}_uploader_version"

    if items_key not in st.session_state:
        st.session_state[items_key] = []

    if uploader_version_key not in st.session_state:
        st.session_state[uploader_version_key] = 0

    st.subheader(title)

    uploaded_files = st.file_uploader(
        f"上传 {title} xlsx 文件，可多选",
        type=["xlsx"],
        accept_multiple_files=True,
        key=f"{state_prefix}_uploader_{st.session_state[uploader_version_key]}"
    )

    # 处理新增上传文件
    if uploaded_files:
        existing_ids = {item["id"] for item in st.session_state[items_key]}

        for uploaded_file in uploaded_files:
            file_bytes = uploaded_file.getvalue()
            file_id = get_uploaded_file_id(uploaded_file.name, file_bytes)

            if file_id not in existing_ids:
                schema_text = read_table_schema_xlsx(file_bytes, uploaded_file.name)

                st.session_state[items_key].append(
                    {
                        "id": file_id,
                        "name": uploaded_file.name,
                        "bytes": file_bytes,
                        "schema_text": schema_text,
                        "partition": ""
                    }
                )

                existing_ids.add(file_id)

    if not st.session_state[items_key]:
        st.info(f"暂未上传{title}文件。")
        return ""

    st.write(f"已上传 {len(st.session_state[items_key])} 个{title}文件：")

    # 渲染已上传文件列表
    for index, item in enumerate(list(st.session_state[items_key]), start=1):
        with st.container(border=True):
            col1, col2 = st.columns([5, 1])

            with col1:
                st.markdown(f"**{index}. {item['name']}**")

            with col2:
                delete_clicked = st.button(
                    "删除",
                    key=f"{state_prefix}_delete_{item['id']}"
                )

            if delete_clicked:
                st.session_state[items_key] = [
                    x for x in st.session_state[items_key]
                    if x["id"] != item["id"]
                ]

                # 更新 uploader key，清空 file_uploader 当前缓存，避免删除后又被自动加回来
                st.session_state[uploader_version_key] += 1
                st.rerun()

            partition_value = st.text_input(
                "分区，可不填；不填表示无分区",
                value=item.get("partition", ""),
                key=f"{state_prefix}_partition_{item['id']}",
                placeholder="例如：pt='20250101'，dt=${bizdate}，ds=2025-01-01"
            )

            # 更新分区
            for saved_item in st.session_state[items_key]:
                if saved_item["id"] == item["id"]:
                    saved_item["partition"] = partition_value.strip()
                    break

            # 注意：这里不能用 st.expander，因为外层已经在 expander 里
            show_preview = st.checkbox(
                "显示解析后的表结构内容",
                key=f"{state_prefix}_show_preview_{item['id']}"
            )

            if show_preview:
                preview_text = item.get("schema_text", "")

                if len(preview_text) > 8000:
                    preview_text = preview_text[:8000] + "\n\n......内容过长，已截断预览......"

                st.text_area(
                    "解析内容预览",
                    value=preview_text,
                    height=260,
                    disabled=True,
                    key=f"{state_prefix}_preview_text_{item['id']}"
                )

    # 拼接最终传给大模型的内容
    final_parts = []

    for item in st.session_state[items_key]:
        partition = item.get("partition", "").strip()
        partition_text = partition if partition else "无分区"

        final_parts.append(
            f"""
### 表结构文件：{item['name']}

分区信息：{partition_text}

表结构内容：
{item.get("schema_text", "")}
"""
        )

    return "\n\n".join(final_parts)


# =========================
# 7. Streamlit 页面配置
# =========================

st.set_page_config(
    page_title="PRD 测试用例生成",
    page_icon=None,
    layout="wide"
)

inject_custom_css()

st.markdown(
    """
<div class="app-hero">
    <div class="app-hero-title">PRD 测试用例生成工具</div>
    <div class="app-hero-desc">
        从 PRD、会议纪要、表结构和开发代码中提炼数据测试需求，自动生成测试用例与 SQL 校验脚本。
    </div>
</div>
    """,
    unsafe_allow_html=True
)
render_step_progress()

# =========================
# 8. 初始化 Session State
# =========================
if "current_step" not in st.session_state:
    st.session_state["current_step"] = STEP_INPUT

if "prd_draft_analysis_result" not in st.session_state:
    st.session_state["prd_draft_analysis_result"] = ""

if "prd_pending_answers" not in st.session_state:
    st.session_state["prd_pending_answers"] = ""

if "prd_final_analysis_result" not in st.session_state:
    st.session_state["prd_final_analysis_result"] = ""

if "test_case_result" not in st.session_state:
    st.session_state["test_case_result"] = ""

if "uploaded_prd_text" not in st.session_state:
    st.session_state["uploaded_prd_text"] = ""

if "prd_file_uploader_version" not in st.session_state:
    st.session_state["prd_file_uploader_version"] = 0

if "source_schema_items" not in st.session_state:
    st.session_state["source_schema_items"] = []

if "result_schema_items" not in st.session_state:
    st.session_state["result_schema_items"] = []

if "source_schema_uploader_version" not in st.session_state:
    st.session_state["source_schema_uploader_version"] = 0

if "result_schema_uploader_version" not in st.session_state:
    st.session_state["result_schema_uploader_version"] = 0

if "pending_points_rows" not in st.session_state:
    st.session_state["pending_points_rows"] = []
if "pending_points_editor_version" not in st.session_state:
    st.session_state["pending_points_editor_version"] = 0

if "prd_current_analysis_result" not in st.session_state:
    st.session_state["prd_current_analysis_result"] = ""

if "pending_analysis_round" not in st.session_state:
    st.session_state["pending_analysis_round"] = 0

if "pending_confirm_history" not in st.session_state:
    st.session_state["pending_confirm_history"] = ""

if "ignore_remaining_pending_points" not in st.session_state:
    st.session_state["ignore_remaining_pending_points"] = False

if "ignored_pending_points_text" not in st.session_state:
    st.session_state["ignored_pending_points_text"] = ""

material_state_defaults = {
    "prd_text": "",
    "prd_manual_text": "",
    "meeting_notes": "",
    "result_table_schema": "",
    "source_table_schema": "",
    "dev_code": "",
}
for key, default_value in material_state_defaults.items():
    if key not in st.session_state:
        st.session_state[key] = default_value


# =========================
# 9. 页面侧边栏
# =========================

with st.sidebar:
    # 模型配置折叠起来
    with st.expander("模型配置", expanded=False):
        st.caption("每个用户填写自己的 API Key，互不影响。配置仅保存在浏览器当前会话中。")

        # 初始化 session_state（用 .env 作为默认值）
        if "user_openai_api_key" not in st.session_state:
            st.session_state["user_openai_api_key"] = _ENV_API_KEY
        if "user_openai_base_url" not in st.session_state:
            st.session_state["user_openai_base_url"] = _ENV_BASE_URL
        if "user_openai_model" not in st.session_state:
            st.session_state["user_openai_model"] = _ENV_MODEL

        st.text_input(
            "API Key",
            key="user_openai_api_key",
            type="password",
            placeholder="sk-xxxx",
            help="你的 LLM API Key，例如 sk-xxx"
        )
        st.text_input(
            "Base URL",
            key="user_openai_base_url",
            placeholder="http://xxx/v1",
            help="LLM 接口地址，注意以 /v1 结尾"
        )
        st.text_input(
            "Model",
            key="user_openai_model",
            placeholder="gpt-4o-mini",
            help="模型名称"
        )

        configured = bool(st.session_state.get("user_openai_api_key", "").strip())
        if configured:
            st.success("API Key 已配置")
        else:
            st.warning("请填写 API Key")

    # 步骤导航放在中心位置
    st.header("步骤导航")
    current_step_value = st.session_state.get("current_step", STEP_INPUT)
    if current_step_value not in STEP_OPTIONS:
        current_step_value = STEP_INPUT
        st.session_state["current_step"] = STEP_INPUT
    # 如果是代码自动跳转，例如 go_to_step(STEP_PENDING)，这里同步侧边栏 radio 的显示值
    if st.session_state.get("current_step_radio") != current_step_value:
        st.session_state["current_step_radio"] = current_step_value
    st.radio(
        "选择当前步骤",
        STEP_OPTIONS,
        label_visibility="collapsed",
        key="current_step_radio",
        on_change=on_step_radio_change
    )

    st.divider()

    if st.button("清空全部结果", use_container_width=True):
        st.session_state["prd_draft_analysis_result"] = ""
        st.session_state["prd_pending_answers"] = ""
        st.session_state["prd_final_analysis_result"] = ""
        st.session_state["test_case_result"] = ""
        st.session_state["uploaded_prd_text"] = ""
        st.session_state["prd_file_uploader_version"] += 1
        st.session_state["source_schema_items"] = []
        st.session_state["result_schema_items"] = []
        st.session_state["source_schema_uploader_version"] += 1
        st.session_state["result_schema_uploader_version"] += 1
        st.session_state["pending_points_rows"] = []
        st.session_state["pending_points_editor_version"] += 1
        st.session_state["prd_current_analysis_result"] = ""
        st.session_state["pending_analysis_round"] = 0
        st.session_state["pending_confirm_history"] = ""
        st.session_state["ignore_remaining_pending_points"] = False
        st.session_state["ignored_pending_points_text"] = ""
        st.session_state["prd_text"] = ""
        st.session_state["prd_manual_text"] = ""
        st.session_state["meeting_notes"] = ""
        st.session_state["result_table_schema"] = ""
        st.session_state["source_table_schema"] = ""
        st.session_state["dev_code"] = ""
        st.session_state["current_step"] = STEP_INPUT
        st.rerun()

# =========================
# 10. 第 1 步：PRD 输入区 + 补充信息区
# =========================

if st.session_state["current_step"] == STEP_INPUT:
    render_page_header(
        title="1. 输入材料",
        desc="上传或粘贴 PRD，并补充会议纪要、源表表结构、结果表表结构、分区信息和开发代码。",
        icon=""
    )

    # =========================
    # 10. PRD 输入区
    # =========================

    st.subheader("一、输入 PRD")

    col_upload, col_manual = st.columns([1, 2])

    with col_upload:
        with st.container(border=True):
            st.markdown("#### 上传 PRD 文件")

            prd_file = st.file_uploader(
                "支持 txt、md、pdf、docx、xlsx、sql、csv、json、py",
                type=["txt", "md", "pdf", "docx", "xlsx", "sql", "csv", "json", "py"],
                key=f"prd_file_{st.session_state['prd_file_uploader_version']}"
            )

            if prd_file is not None:
                uploaded_text = read_uploaded_file(prd_file)
                st.session_state["uploaded_prd_text"] = uploaded_text

            if st.session_state.get("uploaded_prd_text", ""):
                st.success("已读取上传的 PRD 文件。")

                if st.button(
                    "清除已上传内容",
                    use_container_width=True
                ):
                    st.session_state["uploaded_prd_text"] = ""
                    st.session_state["prd_file_uploader_version"] += 1
                    st.rerun()
            else:
                st.info("未上传 PRD 文件，可直接在右侧粘贴内容。")

    with col_manual:
        with st.container(border=True):
            st.markdown("#### 粘贴 PRD 内容")

            prd_manual_text = st.text_area(
                "粘贴 PRD 内容",
                key="prd_manual_text",
                height=260,
                placeholder="请在这里粘贴 PRD 文本。如果已经上传文件，也可以在这里补充说明。",
                label_visibility="collapsed"
            )

    prd_text = ""

    if st.session_state.get("uploaded_prd_text", ""):
        prd_text += st.session_state["uploaded_prd_text"]

    if prd_manual_text.strip():
        prd_text += "\n\n【用户手动补充 PRD 内容】\n" + prd_manual_text

    st.session_state["prd_text"] = prd_text

    if prd_text.strip():
        with st.expander("查看当前 PRD 原文", expanded=False):
            st.text_area(
                "当前合并后的 PRD 内容",
                value=prd_text,
                height=300,
                disabled=True
            )

    # =========================
    # 11. 补充信息区
    # =========================

    st.header("二、补充信息，可选")

    with st.expander("填写会议纪要、表结构、分区、开发代码等补充信息", expanded=False):
        meeting_notes = st.text_area(
            "会议纪要，可选",
            key="meeting_notes",
            height=140,
            placeholder="例如：会议中确认了统计口径、过滤条件、字段含义等。"
        )


        st.divider()

        result_table_schema = render_table_schema_uploader(
            title="结果表表结构",
            state_prefix="result_schema"
        )

        st.session_state["result_table_schema"] = result_table_schema

        st.divider()

        source_table_schema = render_table_schema_uploader(
            title="源表表结构",
            state_prefix="source_schema"
        )

        st.session_state["source_table_schema"] = source_table_schema

        st.divider()

        dev_code = st.text_area(
            "参考开发代码，可选",
            key="dev_code",
            height=220,
            placeholder="可以粘贴 SQL、PySpark、DataWorks 调度代码等。"
        )

        


    # =========================
    # 12. 第一步：生成初版需求提炼和待确认点
    # =========================

    st.header("三、第一步：生成初版需求提炼和待确认点")

    st.info(
        "第一步会根据 PRD、会议纪要、表结构、分区信息、开发代码进行初版分析；不确定的内容会放到待确认点中。"
    )

    if st.button(
        "🚀 生成初版需求提炼和待确认点",
        type="primary",
        use_container_width=True
    ):
        materials = get_materials_from_state()

        if not materials["prd_text"].strip():
            st.warning("请先上传或粘贴 PRD 内容。")
        else:
            with st.spinner("正在分析 PRD..."):
                user_content = f"""
以下是 PRD 原文：

{materials["prd_text"]}

以下是会议纪要：

{materials["meeting_notes"] if materials["meeting_notes"].strip() else "未提供"}

以下是结果表表结构及分区信息：

{materials["result_table_schema"] if materials["result_table_schema"].strip() else "未上传结果表表结构。"}

以下是源表表结构及分区信息：

{materials["source_table_schema"] if materials["source_table_schema"].strip() else "未上传源表表结构。"}

以下是开发代码：

{materials["dev_code"] if materials["dev_code"].strip() else "未提供"}

请基于以上全部信息进行第一轮分析。

请特别注意：
1. 测试用例需要精简，但不能把所有字段的一致性比对强行混在一个用例里。
2. 最重要的测试是“主键唯一性校验”和“结果表与源表加工结果一致性比对”。
3. 字段一致性比对需要按来源表、加工逻辑、过滤条件、关联条件、分区条件和复杂度合理拆分。
4. 简单同源、同关联、同过滤、同分区、直接映射的字段，可以合并到同一条一致性比对用例。
5. 复杂字段、金额字段、枚举映射字段、case when 字段、聚合字段、去重字段、不同源表字段，应单独生成一致性比对用例和 SQL。
6. 每条一致性比对 SQL 的 final select 必须同时展示源表加工后的字段值和结果表字段值。
7. 如果存在差异，SQL 结果应能直观看到 expected/source 值、actual/result 值、diff_flag 和 diff_type。
8. 不要只输出差异数量，要输出差异明细。
9. 复杂字段的 expected CTE 中要保留必要中间字段，方便排查差异原因。
"""

                draft_result = call_llm(
                    PRD_DRAFT_ANALYSIS_PROMPT,
                    user_content
                )
                if is_llm_error(draft_result):
                    st.error(draft_result)
                    st.stop()

                st.session_state["prd_draft_analysis_result"] = draft_result
                st.session_state["prd_current_analysis_result"] = draft_result

                st.session_state["pending_points_rows"] = parse_pending_points_from_markdown(
                    draft_result
                )

                st.session_state["pending_points_editor_version"] += 1

                st.session_state["prd_pending_answers"] = pending_points_to_llm_text(
                    st.session_state["pending_points_rows"]
                )

                st.session_state["pending_analysis_round"] = 1
                st.session_state["pending_confirm_history"] = ""

                st.session_state["ignore_remaining_pending_points"] = False
                st.session_state["ignored_pending_points_text"] = ""

                st.session_state["prd_final_analysis_result"] = ""
                st.session_state["test_case_result"] = ""

            go_to_step(STEP_PENDING)


# =========================
# 13. 第 2 步：展示当前需求分析结果 + 待确认点多轮收敛
# =========================

elif st.session_state["current_step"] == STEP_PENDING:
    render_page_header(
        title="2. 待确认点收敛",
        desc="查看 AI 初版需求分析结果，并在待确认点清单中补充说明，支持多轮收敛。",
        icon=""
    )

    if not st.session_state.get("prd_current_analysis_result"):
        st.warning("请先完成第 1 步：输入材料并生成初版需求分析。")

        if st.button("返回第 1 步", use_container_width=True):
            go_to_step(STEP_INPUT)

        st.stop()

    materials = get_materials_from_state()

    st.subheader("当前需求分析结果")

    st.caption(
        f"当前待确认点解析轮次：第 {st.session_state.get('pending_analysis_round', 1)} 轮"
    )

    current_without_pending_points = remove_pending_points_section(
        st.session_state["prd_current_analysis_result"]
    )

    if st.session_state["pending_points_rows"]:
        if current_without_pending_points.strip():
            render_markdown_in_scroll_box(
                title="查看当前需求分析结果",
                markdown_text=current_without_pending_points,
                height=520,
                expanded=True
            )
    else:
        render_markdown_in_scroll_box(
            title="查看当前完整需求分析结果",
            markdown_text=st.session_state["prd_current_analysis_result"],
            height=520,
            expanded=True
        )

    st.subheader("待确认点清单")

    st.caption(
        "请直接在最后一列“用户补充说明”中填写确认结果。填写后可点击“提交补充说明，继续解析待确认点”。"
    )

    if not st.session_state["pending_points_rows"]:
        st.success("当前没有阻塞测试设计或 SQL 校验的待确认点，可以继续生成最终版需求提炼表。")
        st.session_state["prd_pending_answers"] = "无待确认点。"

        if st.button(
            "进入第 3 步：生成最终版",
            type="primary",
            use_container_width=True
        ):
            go_to_step(STEP_FINAL)

    else:
        render_pending_points_data_editor()

        st.divider()

        col_a, col_b = st.columns([1, 1])

        with col_a:
            continue_pending_clicked = st.button(
                "🔄 提交补充说明，继续收敛",
                type="primary",
                use_container_width=True
            )

        with col_b:
            ignore_pending_clicked = st.button(
                "⏭️ 忽略剩余待确认点，继续",
                use_container_width=True
            )

        if continue_pending_clicked:
            sync_pending_points_from_widgets()

            with st.spinner("正在重新解析..."):
                current_answers = st.session_state["prd_pending_answers"]

                user_content = f"""
以下是 PRD 原文：

{materials["prd_text"]}

以下是上一轮 AI 需求分析结果：

{st.session_state["prd_current_analysis_result"]}

以下是本轮用户针对待确认点填写的补充说明：

{current_answers if current_answers.strip() else "用户未补充。"}

以下是历轮待确认点处理记录：

{st.session_state["pending_confirm_history"] if st.session_state["pending_confirm_history"].strip() else "暂无历史记录。"}

以下是会议纪要：

{materials["meeting_notes"] if materials["meeting_notes"].strip() else "未提供"}

以下是结果表表结构及分区信息：

{materials["result_table_schema"] if materials["result_table_schema"].strip() else "未上传结果表表结构。"}

以下是源表表结构及分区信息：

{materials["source_table_schema"] if materials["source_table_schema"].strip() else "未上传源表表结构。"}

以下是开发代码：

{materials["dev_code"] if materials["dev_code"].strip() else "未提供"}

请基于以上信息进行新一轮需求解析，并重新输出仍需确认的待确认点清单。
"""

                new_analysis_result = call_llm(
                    PRD_ITERATIVE_PENDING_ANALYSIS_PROMPT,
                    user_content
                )
                if is_llm_error(new_analysis_result):
                    st.error(new_analysis_result)
                    st.stop()

                st.session_state["pending_confirm_history"] += f"""

==============================
第 {st.session_state["pending_analysis_round"]} 轮待确认点处理记录
==============================

【用户补充说明】
{current_answers if current_answers.strip() else "用户未补充。"}

【本轮解析前分析结果】
{st.session_state["prd_current_analysis_result"]}

【本轮解析后结果】
{new_analysis_result}
"""

                st.session_state["prd_current_analysis_result"] = new_analysis_result
                st.session_state["prd_draft_analysis_result"] = new_analysis_result

                st.session_state["pending_points_rows"] = parse_pending_points_from_markdown(
                    new_analysis_result
                )

                st.session_state["pending_points_editor_version"] += 1

                st.session_state["prd_pending_answers"] = pending_points_to_llm_text(
                    st.session_state["pending_points_rows"]
                )

                st.session_state["pending_analysis_round"] += 1

                st.session_state["ignore_remaining_pending_points"] = False
                st.session_state["ignored_pending_points_text"] = ""

                st.session_state["prd_final_analysis_result"] = ""
                st.session_state["test_case_result"] = ""

            st.rerun()

        if ignore_pending_clicked:
            sync_pending_points_from_widgets()

            st.session_state["ignore_remaining_pending_points"] = True

            st.session_state["ignored_pending_points_text"] = pending_points_to_llm_text(
                st.session_state["pending_points_rows"]
            )

            st.success("已标记忽略剩余待确认点，可以继续生成最终版需求提炼表。")

            go_to_step(STEP_FINAL)

    if st.session_state.get("ignore_remaining_pending_points", False):
        st.warning(
            "你已选择忽略剩余待确认点。后续生成最终版和测试用例时，AI 会基于当前信息继续处理；被忽略的问题可能在最终版备注或 SQL TODO 中体现。"
        )

    download_content = ""

    if current_without_pending_points.strip():
        download_content += current_without_pending_points.strip()
    else:
        download_content += st.session_state["prd_current_analysis_result"]

    download_content += "\n\n" + pending_points_to_markdown(
        st.session_state["pending_points_rows"]
    )

    if st.session_state.get("pending_confirm_history", "").strip():
        download_content += "\n\n## 历轮待确认点处理记录\n\n"
        download_content += st.session_state["pending_confirm_history"]

    st.download_button(
        label="下载 PRD 分析结果",
        data=download_content,
        file_name="prd_当前需求分析和待确认点.md",
        mime="text/markdown",
        use_container_width=True
    )


# =========================
# 15 + 16. 第 3 步：生成并展示最终版需求提炼表
# =========================

elif st.session_state["current_step"] == STEP_FINAL:
    render_page_header(
        title="3. 最终版需求提炼",
        desc="基于 PRD、补充说明、表结构、分区信息和待确认点处理记录，生成最终版需求提炼表。",
        icon=""
    )

    if not st.session_state.get("prd_current_analysis_result"):
        st.warning("请先完成第 1 步：输入材料并生成初版需求分析。")

        if st.button("返回第 1 步", use_container_width=True):
            go_to_step(STEP_INPUT)

        st.stop()

    has_pending_points = bool(st.session_state.get("pending_points_rows", []))
    ignored_pending = st.session_state.get("ignore_remaining_pending_points", False)

    if has_pending_points and not ignored_pending:
        st.warning("当前仍存在待确认点。请先在第 2 步补充说明，或者选择忽略剩余待确认点。")

        if st.button("返回第 2 步处理待确认点", use_container_width=True):
            go_to_step(STEP_PENDING)

        st.stop()

    materials = get_materials_from_state()

    if ignored_pending:
        st.info("当前存在被用户选择忽略的待确认点，仍允许生成最终版需求提炼表。")
    else:
        st.success("当前无待确认点，可以生成最终版需求提炼表。")

    if st.button(
        "生成最终版需求提炼表",
        type="primary",
        use_container_width=True
    ):
        with st.spinner("正在生成最终版..."):
            sync_pending_points_from_widgets()

            user_content = f"""
以下是 PRD 原文：

{materials["prd_text"]}

以下是经过多轮待确认点收敛后的最新需求分析结果：

{st.session_state["prd_current_analysis_result"]}

以下是历轮待确认点处理记录：

{st.session_state["pending_confirm_history"] if st.session_state["pending_confirm_history"].strip() else "暂无历史记录。"}

以下是当前待确认点及用户补充说明：

{st.session_state["prd_pending_answers"] if st.session_state["prd_pending_answers"].strip() else "无待确认点。"}

以下是用户选择忽略的剩余待确认点：

{st.session_state["ignored_pending_points_text"] if st.session_state.get("ignore_remaining_pending_points", False) else "无。"}

以下是会议纪要：

{materials["meeting_notes"] if materials["meeting_notes"].strip() else "未提供"}

以下是结果表表结构及分区信息：

{materials["result_table_schema"] if materials["result_table_schema"].strip() else "未上传结果表表结构。"}

以下是源表表结构及分区信息：

{materials["source_table_schema"] if materials["source_table_schema"].strip() else "未上传源表表结构。"}

以下是开发代码：

{materials["dev_code"] if materials["dev_code"].strip() else "未提供"}

请基于以上信息生成最终版需求提炼表。

特别要求：
1. 对已经确认清楚的问题，状态标记为“已明确”。
2. 对用户明确选择忽略的待确认点，不要继续阻塞最终版输出。
3. 如果用户忽略的问题会影响 SQL 准确性，需要在测试关注点、备注或仍需确认问题中标记“用户选择忽略”。
4. 不要再次输出低价值或不影响测试 SQL 的待确认问题。
5. 最终版需求提炼必须能直接用于后续生成数据测试用例和 SQL 校验脚本。
"""

            final_result = call_llm(
                PRD_FINAL_ANALYSIS_PROMPT,
                user_content
            )
            if is_llm_error(final_result):
                st.error(final_result)
                st.stop()

            st.session_state["prd_final_analysis_result"] = final_result
            st.session_state["test_case_result"] = ""

        st.success("最终版需求提炼表已生成。")

    if st.session_state["prd_final_analysis_result"]:
        st.subheader("最终版需求提炼表")

        render_markdown_in_scroll_box(
            title="查看最终版需求提炼表",
            markdown_text=st.session_state["prd_final_analysis_result"],
            height=560,
            expanded=True
        )

        st.download_button(
            label="下载需求提炼表",
            data=st.session_state["prd_final_analysis_result"],
            file_name="prd_最终版需求提炼表.md",
            mime="text/markdown",
            use_container_width=True
        )

        st.divider()

        if st.button(
            "进入第 4 步：生成测试用例",
            type="primary",
            use_container_width=True
        ):
            go_to_step(STEP_TEST_CASE)


# =========================
# 17 + 18. 第 4 步：生成并展示测试用例和 SQL
# =========================

elif st.session_state["current_step"] == STEP_TEST_CASE:
    render_page_header(
        title="4. 测试用例与 SQL",
        desc="根据最终版需求提炼表生成精简、可定位问题的数据测试用例和 SQL 校验脚本。",
        icon=""
    )

    if not st.session_state.get("prd_final_analysis_result"):
        st.warning("请先完成第 3 步：生成最终版需求提炼表。")

        if st.button("返回第 3 步", use_container_width=True):
            go_to_step(STEP_FINAL)

        st.stop()

    materials = get_materials_from_state()

    if st.button(
        "生成测试用例和 SQL",
        type="primary",
        use_container_width=True
    ):
        with st.spinner("正在生成测试用例..."):
            sync_pending_points_from_widgets()

            user_content = f"""
以下是 PRD 原文：

{materials["prd_text"]}

以下是最终版需求提炼表：

{st.session_state["prd_final_analysis_result"]}

以下是历轮待确认点处理记录：

{st.session_state["pending_confirm_history"] if st.session_state["pending_confirm_history"].strip() else "暂无历史记录。"}

以下是当前待确认点及用户补充说明：

{st.session_state["prd_pending_answers"] if st.session_state["prd_pending_answers"].strip() else "无待确认点。"}

以下是用户选择忽略的剩余待确认点：

{st.session_state["ignored_pending_points_text"] if st.session_state.get("ignore_remaining_pending_points", False) else "无。"}

以下是补充信息。补充信息均为可选，如果没有提供，请不要自行脑补。

【会议纪要】
{materials["meeting_notes"] if materials["meeting_notes"].strip() else "未提供"}

【结果表表结构 / 字段说明 / 分区信息】
{materials["result_table_schema"] if materials["result_table_schema"].strip() else "未上传结果表表结构。"}

【源表表结构 / 字段说明 / 分区信息】
{materials["source_table_schema"] if materials["source_table_schema"].strip() else "未上传源表表结构。"}

【参考开发代码】
{materials["dev_code"] if materials["dev_code"].strip() else "未提供"}

请基于以上信息生成数据测试用例和 SQL 校验脚本。

要求：
1. 必须以“最终版需求提炼表”为主要依据。
2. 如果最终版需求提炼表中仍存在“待确认”“部分待确认”“存在冲突”的内容，需要在测试用例中标记。
3. 如果存在“用户选择忽略”的待确认点，需要在测试用例前置条件、SQL 注释或待确认问题清单中标记。
4. 不允许根据初版分析中的不确定内容自行脑补。
5. 如果源表或结果表分区未提供，请在 SQL 中使用【分区字段】或 pt='YYYYMMDD' 占位，并在待确认问题中说明。
6. 如果某个上传表结构的分区信息为“无分区”，不要强行给该表添加分区条件。
"""

            test_result = call_llm(
                TEST_GEN_PROMPT,
                user_content
            )
            if is_llm_error(test_result):
                st.error(test_result)
                st.stop()

            st.session_state["test_case_result"] = test_result

        st.success("测试用例和 SQL 校验脚本已生成。")

    if st.session_state["test_case_result"]:
        st.subheader("测试用例和 SQL 校验脚本")

        render_test_case_result_with_download(
            st.session_state["test_case_result"]
        )
else:
    st.warning("当前步骤状态异常，已返回第 1 步。")
    st.session_state["current_step"] = STEP_INPUT
    st.rerun()
