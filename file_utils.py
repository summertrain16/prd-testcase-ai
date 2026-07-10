"""
文件读取工具方法。

包含：
- read_xlsx_file: 读取普通 xlsx 文件内容，转换成文本
- read_uploaded_file: 支持读取 txt、md、sql、csv、json、py、pdf、docx、xlsx
- read_table_schema_xlsx: 读取表结构 xlsx 文件内容
- read_html_url: 从 URL 抓取 HTML 页面内容，转成纯文本
- get_uploaded_file_id: 根据文件名和文件内容生成唯一 ID
"""

import hashlib
from io import BytesIO

import pandas as pd
from pypdf import PdfReader
from docx import Document
from openpyxl import load_workbook


def read_xlsx_file(uploaded_file) -> str:
    """
    读取普通 xlsx 文件内容，并转换成文本，方便大模型分析。
    用于 PRD Excel 文件读取。
    """
    try:
        uploaded_file.seek(0)

        workbook = load_workbook(
            uploaded_file,
            data_only=True,
            read_only=True
        )

        result = []

        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            result.append(f"\n\n## Excel Sheet：{sheet_name}\n")

            rows = list(sheet.iter_rows(values_only=True))

            if not rows:
                result.append("该 Sheet 为空。\n")
                continue

            max_rows = 300
            rows = rows[:max_rows]

            for row_index, row in enumerate(rows, start=1):
                values = []
                for cell in row:
                    if cell is None:
                        values.append("")
                    else:
                        values.append(str(cell).strip())
                result.append(f"第 {row_index} 行：" + " | ".join(values))

            if sheet.max_row > max_rows:
                result.append(
                    f"\n注意：该 Sheet 共 {sheet.max_row} 行，当前只读取前 {max_rows} 行。"
                )

        return "\n".join(result)

    except Exception as e:
        return f"xlsx 文件读取失败：{str(e)}"


def read_uploaded_file(uploaded_file) -> str:
    """
    支持读取 txt、md、sql、csv、json、py、pdf、docx、xlsx。
    用于 PRD 文件读取。
    """
    if uploaded_file is None:
        return ""

    file_name = uploaded_file.name.lower()

    try:
        if file_name.endswith((".txt", ".md", ".sql", ".csv", ".json", ".py")):
            uploaded_file.seek(0)
            return uploaded_file.read().decode("utf-8", errors="ignore")

        elif file_name.endswith(".pdf"):
            uploaded_file.seek(0)
            reader = PdfReader(uploaded_file)
            text = ""
            for page in reader.pages:
                page_text = page.extract_text()
                if page_text:
                    text += page_text + "\n"
            return text

        elif file_name.endswith(".docx"):
            uploaded_file.seek(0)
            doc = Document(uploaded_file)
            text = ""
            for para in doc.paragraphs:
                text += para.text + "\n"
            return text

        elif file_name.endswith(".xlsx"):
            return read_xlsx_file(uploaded_file)

        else:
            return "暂不支持该文件格式，请上传 txt、md、sql、csv、json、py、pdf、docx 或 xlsx 文件。"

    except Exception as e:
        return f"文件读取失败：{str(e)}"


def read_html_url(url: str) -> str:
    """
    从 URL 抓取 HTML 页面内容，提取正文文本。

    使用 requests + BeautifulSoup 去除 script/style/nav 等噪音标签，
    保留正文内容转成纯文本，用于 PRD 分析。

    参数：
        url: 页面 URL，例如 https://xxx.app.codebuddy.work/

    返回：
        提取后的纯文本。失败返回错误信息字符串。
    """
    try:
        import requests
        from bs4 import BeautifulSoup

        _headers = {
            "User-Agent": (
                "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                "AppleWebKit/537.36 (KHTML, like Gecko) "
                "Chrome/125.0.0.0 Safari/537.36"
            )
        }

        resp = requests.get(url, headers=_headers, timeout=20)
        resp.raise_for_status()
        resp.encoding = resp.apparent_encoding or "utf-8"

        soup = BeautifulSoup(resp.text, "html.parser")

        # 去除噪音标签
        for tag in soup(["script", "style", "nav", "footer", "header", "noscript", "svg"]):
            tag.decompose()

        # 提取标题
        title = soup.title.get_text(strip=True) if soup.title else ""

        # 提取正文
        body = soup.body or soup
        text = body.get_text(separator="\n", strip=True)

        # 合并标题
        parts = []
        if title:
            parts.append(f"# {title}\n")
        parts.append(text)

        result = "\n".join(parts)

        # 截断超长内容（最大 10 万字符）
        if len(result) > 100000:
            result = result[:100000] + "\n\n......内容过长，已截断......"

        return result

    except Exception as e:
        return f"URL 页面抓取失败：{str(e)}"


def read_table_schema_xlsx(file_bytes: bytes, file_name: str) -> str:
    """
    读取表结构 xlsx 文件内容，转换成文本。
    一个 xlsx 可以包含多个 sheet。
    用于源表/结果表结构上传。
    """
    try:
        workbook = load_workbook(
            BytesIO(file_bytes),
            data_only=True,
            read_only=True
        )

        result = []

        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]

            result.append(f"\n## 文件：{file_name}")
            result.append(f"## Sheet：{sheet_name}")

            rows = []

            for row in sheet.iter_rows(values_only=True):
                values = []
                has_value = False

                for cell in row:
                    if cell is None:
                        values.append("")
                    else:
                        cell_value = str(cell).strip()
                        values.append(cell_value)
                        if cell_value:
                            has_value = True

                if has_value:
                    rows.append(values)

            if not rows:
                result.append("该 Sheet 为空。")
                continue

            max_rows = 1000
            original_row_count = len(rows)
            rows = rows[:max_rows]

            for index, row in enumerate(rows, start=1):
                result.append(f"第 {index} 行：" + " | ".join(row))

            if original_row_count > max_rows:
                result.append(
                    f"注意：该 Sheet 共 {original_row_count} 行，当前只读取前 {max_rows} 行。"
                )

        return "\n".join(result)

    except Exception as e:
        return f"xlsx 表结构文件读取失败：{str(e)}"


def get_uploaded_file_id(file_name: str, file_bytes: bytes) -> str:
    """
    根据文件名和文件内容生成唯一 ID，用于去重和删除。
    """
    md5 = hashlib.md5()
    md5.update(file_name.encode("utf-8"))
    md5.update(file_bytes)
    return md5.hexdigest()
